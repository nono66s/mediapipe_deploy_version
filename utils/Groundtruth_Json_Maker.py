import os 
import json
import datetime
import glob
import openpyxl

from utils.Config import config 

class Dir2JSON:
    def __init__(self,input_path):
        
        self.dataset_path   =   input_path

        #------------------------------------#
        #  Make Ground Truth File
        #  先建立Json檔案
        #  內容再透過其他方法寫入
        #------------------------------------#

        #------------------------------------#
        #  獲取時間
        #------------------------------------#
        now             = datetime.datetime.now()
        record_time     = now.strftime('%Y-%m-%d %H:%M:%S.%f')
        file_time       = now.strftime("%Y%m%d-%H%M%S")

        #------------------------------------#
        #  建立Json檔案模板 
        #  版權訊息
        #  資料集名稱
        #  創建該檔案的時間
        #  所有的類別，與該類別所在的資料夾 
        #------------------------------------#
        dataset = dict(
                        license         =   dict(
                                                    __right__     = 'All rights reserved',
                                                    __author__    = "Nono Ju",
                                                    __copyright__ = 'Copyright by Nono Ju ',
                                                    __license__   = "GPL",
                                                    __version__   = "None",
                                                    __email__     = "d610112005@tmu.edu.tw",
                                                    __status__    = 'Working',                          #"Production"
                                                    __inherit__   = 'Mask RCNN json_to_dataset_v6',
                        ),
                        dataset_name     =   config.Dataset_name,
                        time             =   record_time,
                        datasets_contain =   [],
                        categories       =   {},
                        datas_name       =   [],
                        datas            =   {},
                        ground_truth     =   {},
        )
        

        dataset['datasets_contain'].append(self.dataset_path)   
 

        name           = self.get_name(self.xlsx_path_name(),[5,6,7,8,19,20,21,22,25,26,27,28,34,35,36,37,40,41,42,43,49,50,51,52,55,56,57,58,64,65,66,67,70,71,72,73])   #　interval　5,118　
        file_name,data = self.get_data(self.xlsx_path_name(),[5,6,7,8,19,20,21,22,25,26,27,28,34,35,36,37,40,41,42,43,49,50,51,52,55,56,57,58,64,65,66,67,70,71,72,73])   # 5 6 7 8 21 22  27 28 

        dataset['datas_name'].append(name)
        dataset['datas'][self.output_path_name()] = data

        with open(self.output_path_name(), 'w', encoding="utf-8") as f:   
            json.dump(
                dataset, 
                f, 
                indent= 4,
                ensure_ascii=False
                )
            
        return 

    def output_path_name(self):

        file_name=self.dataset_path.split("\\")[-1].split(".")[0]
        
        return str(config.output_path+file_name+"_input.json")

    def xlsx_path_name(self):

        file_name=self.dataset_path.split("\\")[-1].split(".")[0]
        
        return str(config.output_path+file_name+"_excel.xlsx")


    def get_data(self,path_name,interval):
        
        name=path_name.rsplit("_", maxsplit=2)[0]

        wb = openpyxl.load_workbook(path_name, data_only=True) 

        s1 = wb["Sheet1"]

        #tmp計算stdev
        tmp_all=[]

        for j in interval: # 26,56,2     target_columns

            #每個column的開始與結束
            start_num=503
            #end_num=s1.max_row
            end_num=s1.max_row

            tmp=[]

            for k in range(start_num,end_num):
                print(s1.cell(k, j).value)
                tmp.append(s1.cell(k, j).value)

            tmp_all.append(tmp)

        return name,tmp_all
    
    def get_name(self,path_name,interval): 

        wb = openpyxl.load_workbook(path_name, data_only=True) 

        s1 = wb["Sheet1"]

        tmp=[]

        for j in interval: # 26,56,2     target_columns
            tmp.append(s1.cell(2, j).value)

        return tmp



